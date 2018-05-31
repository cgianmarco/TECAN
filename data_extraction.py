from openpyxl import load_workbook
import numpy as np
import os


filename = "/home/gianmarco/Desktop/File Tecan/Growth_Lux/LAB 3.2 - Growth_Lux_13 h_20170628_093816.xlsx"


class FileProcessor():

	def __init__(self, doc):
		self.doc = doc	

	def filter_empty(self, value):
		if value == '':
			value = 0.0
		return value

	def row_of(self, cell):
		return int(filter(lambda x: x.isdigit(), cell))

	def cells_of_value(self, value, start = 1, end = None):
		line = start
		label = 'A' + str(line)
		max_row = self.doc.max_row

		if not end:
			end = max_row

		while line < end:
			if self.doc[label].value == value:
				yield label
			label = label.replace(str(line), str(line+1))
			line += 1

	def cell_of_value(self, value):
		line = 1
		label = 'A' + str(line)
		max_row = self.doc.max_row

		for line in range(1, max_row):
			if self.doc[label].value == value:
				return label
			label = label.replace(str(line), str(line+1))
			line += 1

	def matrix_at(self, line):
		result = []
		for line in range(line + 1, line + 9):
			row = list(self.doc[line])
			result.append([ self.filter_empty(cell.value) for cell in row[1:13] ])
		return result


	def analysis_names(self):
		start = self.row_of(self.cell_of_value("List of actions in this measurement script:"))
		end = self.row_of(self.cell_of_value("Name"))

		for row in range(start, end):
			if self.doc['G' + str(row)].value is not None:
				yield self.doc['G' + str(row)].value

	def tensor_info(self, names):

		if len(names) > 1:
			for first, second in zip(names, names[1:]):
				start = self.row_of(self.cell_of_value(first))
				end = self.row_of(self.cell_of_value(second))
				yield {'name' : first, 'start' : start, 'end' : end}

			yield {'name' : second, 'start' : end, 'end' : None}
			
		else:

			yield {'name': names[0], 'start': 1, 'end': None}



	def tensors(self):
		names = [ name for name in self.analysis_names() ]
		print(names)

		for info in self.tensor_info(names):
			matrices = self.cells_of_value('<>', start = info['start'], end = info['end'])
			tensor = np.asarray([ self.matrix_at(self.row_of(cell)) for cell in matrices])
			yield { 'name' : info['name'], 'data' : tensor }



# def test():
# 	folder = "/home/gianmarco/Desktop/File Tecan/Growth_Lux"
# 	folder = "/home/gianmarco/Desktop/File Tecan/OD600 Single"
# 	global doc

# 	for filename in os.listdir(folder):
# 		print(filename)		
# 		doc = doc_from(os.path.join(folder, filename))
# 		tensor = tensors()
# 		print(next(tensor))

# test()







