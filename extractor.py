from openpyxl import load_workbook, Workbook
import os
from templates import *

filename = "/home/gianmarco/Desktop/File Tecan/Growth_Lux/LAB 3.2 - Growth_Lux_13 h_20170628_093816.xlsx"
OD_SINGLE = "/home/gianmarco/Desktop/Tirocinio/Letture TECAN/OD600/LID/50 ul.xlsx"
OD_SPECTRUM = "/home/gianmarco/Desktop/Tirocinio/Letture TECAN/Spettro OD550-650/LID/50 ul.xlsx"
MULTIPLE = "/home/gianmarco/Desktop/File Tecan/Growth_Lux/LAB 3.2 - Growth_Lux_13 h_20170628_155630.xlsx"
OD_TIME = "/home/gianmarco/Desktop/Excel Tecan/prova.xlsx"
OD_TIME_SPECTRUM = OD_TIME
OD_TIMESPECTRUM_TEMPLATE_TEST = "/home/gianmarco/Desktop/Book2.xlsx"

class Document:
	def __init__(self, worksheet):
		self.doc = worksheet

	def get_dims(self, data_starting_line):
		width = len([cell for cell in self.doc[data_starting_line][1:] if 'A' in cell.value])
		height = len(self.doc[data_starting_line][1:])/width
		return width, height

	def filter_empty(self, value):
			if value == '':
				return 0.0
			else:
				return value

	def cells_of_value(self, value, start = 1, end = None):
		line = start
		label = 'A' + str(line)
		max_row = self.doc.max_row
		labels = []

		if not end:
			end = max_row

		while line < end:
			if self.doc[label].value == value:
				labels.append(label)
			label = label.replace(str(line), str(line+1))
			line += 1

		return labels

	def get_two_dim_matrix(self, data_starting_line):
		result = []
		for line in range(data_starting_line + 1, data_starting_line + 9):
			row = list(self.doc[line])
			result.append([ self.filter_empty(cell.value) for cell in row[1:] ])
		return result

	def get_two_dim_time_spectrum_matrix(self, data_starting_line):
		result = []
		line = data_starting_line
		while self.doc['A' + str(line)].value is not None:
			# print(line)
			if str(self.doc['A' + str(line)].value.encode('utf-8')).isdigit():
				result.append([cell.value for cell in self.doc[line][1:]])
			line += 1
		return result



	def row_of(self, cell):
		return int(''.join(list(filter(lambda x: x.isdigit(), cell))))

	def cell_of_value(self, value):
		line = 1
		label = 'A' + str(line)
		max_row = self.doc.max_row

		for line in range(1, max_row):
			if self.doc[label].value == value:
				return label
			label = label.replace(str(line), str(line+1))
			line += 1








class FileParser:
	def __init__(self, filename):
		filename = filename
		self.wb = load_workbook(str(filename))
		self.doc = self.wb['Result sheet']
		self.document = Document(self.doc)

	def split_worksheet(self, start, end):
		ws_original = self.doc
		new_wb = Workbook()
		ws_old = new_wb.active
		if end == None:
			end = ws_original.max_row
		ws = new_wb.create_sheet("{start}-{end}".format(start=start, end=end))
		new_wb.remove_sheet(ws_old)
		for row in ws_original.iter_rows(min_col=1, min_row=start, max_col=ws_original.max_column, max_row=end):
			for cell in row:
				ws[cell.coordinate] = cell.value

		new_wb.save("{start}-{end}.xlsx".format(start=start, end=end))
		return ws

	


	def tensors(self):
		tensors = []

		# Get names
		names = []

		start = self.document.row_of(self.document.cell_of_value("List of actions in this measurement script:"))
		end = self.document.row_of(self.document.cell_of_value("Name"))

		for row in range(start, end):
			if self.doc['G' + str(row)].value is not None:
				names.append(self.doc['G' + str(row)].value)


		# Find Intervals
		intervals = []

		if len(names) > 1:
			for first, second in zip(names, names[1:]):
				print(names)
				start = self.document.row_of(self.document.cell_of_value(first))
				end = self.document.row_of(self.document.cell_of_value(second))

				interval = (start, end-1)
				intervals.append(interval)

			interval = (end, None)
			intervals.append(interval)
				
		else:
			interval = (1, None)
			intervals.append(interval)


		# Split worksheet
		split_docs = []

		for (start, end) in intervals:
			split_docs.append(Document(self.split_worksheet(start, end)))


		# Extract tensors
		tensors = []

		filetypes = [ODSingle(), ODSpectrum(), ODTime(), ODTimeSpectrum()]
		for doc in split_docs:
			for filetype in filetypes:
				if filetype.test(doc):
					tensors.append(filetype.parse(doc))

		if len(names) > len(tensors):
			names = names[:len(tensors)]

		print(names)

		return names, tensors


# fp = FileParser(OD_TIME)
# print(fp.tensors())


def test():
	doc = Document(load_workbook(OD_TIMESPECTRUM_TEMPLATE_TEST).active)

	print(doc.get_two_dim_time_spectrum_matrix(1))

# test()