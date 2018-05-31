from openpyxl import load_workbook
import numpy as np

def filter_empty(value):
			if value == '':
				return 0.0
			else:
				return value

def get_two_dim_matrix(doc, data_starting_line):
	result = []
	for line in range(data_starting_line + 1, data_starting_line + 9):
		row = list(doc[line])
		result.append([ filter_empty(cell.value) for cell in row[1:] ])
	return result

def get_dims(doc, data_starting_line):
	width = len([cell for cell in doc[data_starting_line][1:] if 'A' in cell.value])
	height = len(doc[data_starting_line][1:])/width
	return width, height

def cell_of_value(value, doc):
		line = 1
		label = 'A' + str(line)
		max_row = doc.max_row

		for line in range(1, max_row):
			if doc[label].value == value:
				return label
			label = label.replace(str(line), str(line+1))
			line += 1

def row_of(cell):
		return int(filter(lambda x: x.isdigit(), cell))





class FileLoader:
	def __init__(self, filename):
		self.filename = filename
		self.doc = load_workbook(str(filename))
		self.filetypes = [ ODSpectrum(), ODSingle(), ODTime(), ODTimeSpectrum()]
		self.doc = self.doc['Result sheet']
		

	def parse(self):
		parsed = []
		for filetype in self.filetypes:
			if filetype.test(self.doc) :
				parsed.append(filetype.parse(self.doc))
		if len(parsed) > 0:
			return parsed
		else:
			raise NotImplementedError('FileType not implemented')



class ODSpectrum():
	def test(self, doc):
		return doc['A24'].value == 'Wavelength start'

	def parse(self, doc):
		data_starting_line = row_of(cell_of_value('Wavel.', doc))
		wl_start = int(str(doc['E24'].value).replace('L', ''))
		wl_end = int(str(doc['E25'].value).replace('L', ''))

		depth = wl_end - wl_start + 1
		width, height = get_dims(doc, data_starting_line)
		time = 1

		# Get values
		result = []
		for i in range(depth):
			row = list(doc[data_starting_line + 1 + i])
			result.append([ cell.value for cell in row[1:] ])

		data = np.transpose(np.array(result).reshape([1, depth, width, height]), (0,1,3,2))

		axis_values = { 'time' : range(time), 
						'depth' : range(wl_start,wl_end), 
						'width' : range(width), 
						'height' : range(height) }

		return {"data" : data, 'axis_values':axis_values}


class Vector():
	def test(self, doc):
		return True

	def parse(self, doc):
		axis_values = []

		depth = 1
		time = 1
		width = 1
		height = 10

		# Get values
		data = np.ones((time, depth, width, height))

	

		axis_values = { 'time' : range(time), 
						'depth' : range(depth), 
						'width' : range(width), 
						'height' : range(height) }


		return {"data" : data, 'axis_values':axis_values}



class ODSingle():
	def test(self, doc):
		return doc['A24'].value == 'Measurement wavelength'

	def parse(self, doc):
		data_starting_line = row_of(cell_of_value('<>', doc))
		axis_values = []

		depth = 1
		time = 1
		width = 12
		height = 8

		# Get values
		result = get_two_dim_matrix(doc, data_starting_line)

	

		axis_values = { 'time' : range(time), 
						'depth' : range(depth), 
						'width' : range(width), 
						'height' : range(height) }

		data = np.array(result).T
		data = np.expand_dims(data, axis=0)
		data = np.expand_dims(data, axis=0)

		return {"data" : data, 'axis_values':axis_values}



class ODTime:

	def test(self, doc):
		print(doc['A106'].value)
		return doc['A106'].value == 'OD600'

	def parse(self, doc):

		# Get values
		result = []
		time_values = []
		matrix = cell_of_values('<>')
		line = row_of(next(matrix))
		while doc[line][0].value == '<>':
			result.append(get_two_dim_matrix(doc, line))
			line += 13



		data = np.array(result)
		data = np.transpose(data, (0,2,1))
		data = np.expand_dims(data, axis=1)

		time, depth, width, height = data.shape
		print(data.shape)

		axis_values = { 'time' : range(time), 
						'depth' : range(depth), 
						'width' : range(width), 
						'height' : range(height) }
		print(axis_values)

		return {"data" : data, 'axis_values':axis_values}


class ODTimeSpectrum:

	def test(self, doc):
		return doc['A316'].value == 'Spettro Eccitazione'

	def parse(self, doc):
		data_starting_line = 321

		# Get values
		result = []
		time_values = []
		line = data_starting_line
		while doc[line][0].value == '400':	
			result.append([[cell.value for cell in doc[row][1:] if cell.value is not None ] for row in range(line, line + 21)])
			line += 26
			print(line)


		data = np.array(result)
		newshape = (8, 12, data.shape[1], data.shape[2])
		data = data.reshape(newshape)
		data = np.transpose(data, (3,2,1,0))

		time, depth, width, height = data.shape

		for t in range(time):
			for z in range(depth):
				for i in range(width):
					for j in range(height):
						if data[t,z,i,j] == 'OVER':
							data[t,z,i,j] = np.nan

		data = data.astype('float32')

		axis_values = { 'time' : range(time), 
						'depth' : range(depth), 
						'width' : range(width), 
						'height' : range(height) }

		return {"data" : data, 'axis_values':axis_values}





	